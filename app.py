import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="SISTEMA DE COBRANZA - RESULTADOS", layout="wide", initial_sidebar_state="expanded")

# Estilos CSS personalizados
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #1f77b4;
    }
    .success-card {
        border-left-color: #28a745;
    }
    .warning-card {
        border-left-color: #ffc107;
    }
    .danger-card {
        border-left-color: #dc3545;
    }
    </style>
""", unsafe_allow_html=True)

st.sidebar.title("üè¢ SISTEMA DE COBRANZA")
st.sidebar.markdown("---")

menu = st.sidebar.radio(
    "üìã MEN√ö PRINCIPAL",
    [
        "üìä Dashboard Cruce Deuda vs Pagos",
        "üì≤ GENERADOR DE SMS",
        "üöß M√≥dulo Hist√≥rico (En Desarrollo)"
    ]
)

# ==========================================================
# MODULO 1 - DASHBOARD PROFESIONAL DE COBRANZA
# ==========================================================

def modulo_cruce():

    st.markdown('<div class="main-header">‚öñÔ∏è DASHBOARD EJECUTIVO DE GESTI√ìN DE COBRANZA</div>', unsafe_allow_html=True)

    def limpiar_columnas(df):
        df.columns = df.columns.str.strip().str.upper().str.replace(" ", "_")
        return df

    # Inicializar session state
    if "df_deuda_base" not in st.session_state:
        st.session_state.df_deuda_base = None

    # ========== CARGA DE CARTERA BASE ==========
    if st.session_state.df_deuda_base is None:

        st.info("üîπ **Paso 1:** Carga la base de CARTERA/DEUDA (se guardar√° en memoria)")
        
        archivo_deuda = st.file_uploader(
            "üìÇ Subir archivo CARTERA / DEUDA",
            type=["xlsx"],
            help="Debe contener las columnas: ID_COBRANZA, PERIODO, DEUDA, TIPO"
        )

        if archivo_deuda:
            
            with st.spinner("Procesando cartera..."):
                try:
                    df_deuda = pd.read_excel(archivo_deuda)
                    df_deuda = limpiar_columnas(df_deuda)

                    columnas_deuda = {"ID_COBRANZA", "PERIODO", "DEUDA", "TIPO"}

                    if not columnas_deuda.issubset(df_deuda.columns):
                        st.error("‚ùå El archivo CARTERA no tiene las columnas obligatorias: ID_COBRANZA, PERIODO, DEUDA, TIPO")
                        return

                    df_deuda["ID_COBRANZA"] = df_deuda["ID_COBRANZA"].astype(str)
                    df_deuda["PERIODO"] = df_deuda["PERIODO"].astype(str)
                    df_deuda["DEUDA"] = pd.to_numeric(df_deuda["DEUDA"], errors="coerce").fillna(0)

                    # Validaciones adicionales
                    if (df_deuda["DEUDA"] < 0).any():
                        st.warning("‚ö†Ô∏è Se detectaron montos negativos en DEUDA. Se convertir√°n a positivos.")
                        df_deuda["DEUDA"] = df_deuda["DEUDA"].abs()

                    st.session_state.df_deuda_base = df_deuda

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("üìÑ Registros Cargados", f"{len(df_deuda):,}")
                    with col2:
                        st.metric("üí∞ Cartera Total", f"Bs. {df_deuda['DEUDA'].sum():,.2f}")
                    with col3:
                        st.metric("üìÖ Periodos", df_deuda["PERIODO"].nunique())

                    st.success("‚úÖ Cartera cargada correctamente y guardada en memoria.")
                    st.balloons()
                    st.rerun()

                except Exception as e:
                    st.error(f"‚ùå Error al procesar el archivo: {str(e)}")
                    return

        return

    else:
        # Cartera ya cargada
        df_deuda = st.session_state.df_deuda_base
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success("‚úÖ **Cartera base cargada en memoria**")
        with col2:
            if st.button("üîÑ Reemplazar Cartera", use_container_width=True):
                st.session_state.df_deuda_base = None
                st.rerun()

        # Mostrar resumen de cartera
        with st.expander("üìä Ver resumen de Cartera Base"):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("üìÑ Total Registros", f"{len(df_deuda):,}")
            with col2:
                st.metric("üí∞ Cartera Total", f"Bs. {df_deuda['DEUDA'].sum():,.2f}")
            with col3:
                st.metric("üìÖ Periodos", df_deuda["PERIODO"].nunique())
            with col4:
                st.metric("üè∑Ô∏è Tipos de Deuda", df_deuda["TIPO"].nunique())

    st.markdown("---")

    # ========== CARGA DE PAGOS ==========
    st.info("üîπ **Paso 2:** Carga el archivo de PAGOS para realizar el cruce")
    
    archivo_pagos = st.file_uploader(
        "üíµ Subir archivo PAGOS (Puede actualizarse constantemente)",
        type=["xlsx"],
        help="Debe contener las columnas: ID_COBRANZA, PERIODO, IMPORTE"
    )

    if not archivo_pagos:
        return

    # Procesar archivo de pagos
    with st.spinner("Procesando pagos y realizando cruce..."):
        try:
            df_deuda = st.session_state.df_deuda_base.copy()
            df_pagos = pd.read_excel(archivo_pagos)
            df_pagos = limpiar_columnas(df_pagos)

            columnas_pagos = {"ID_COBRANZA", "PERIODO", "IMPORTE"}

            if not columnas_pagos.issubset(df_pagos.columns):
                st.error("‚ùå El archivo PAGOS no tiene las columnas obligatorias: ID_COBRANZA, PERIODO, IMPORTE")
                return

            df_pagos["ID_COBRANZA"] = df_pagos["ID_COBRANZA"].astype(str)
            df_pagos["PERIODO"] = df_pagos["PERIODO"].astype(str)
            df_pagos["IMPORTE"] = pd.to_numeric(df_pagos["IMPORTE"], errors="coerce").fillna(0)

            # Validaciones
            if (df_pagos["IMPORTE"] < 0).any():
                st.warning("‚ö†Ô∏è Se detectaron montos negativos en PAGOS. Se convertir√°n a positivos.")
                df_pagos["IMPORTE"] = df_pagos["IMPORTE"].abs()

            # Agrupar pagos
            pagos_resumen = df_pagos.groupby(
                ["ID_COBRANZA", "PERIODO"]
            )["IMPORTE"].sum().reset_index()

            pagos_resumen.rename(columns={"IMPORTE": "TOTAL_PAGADO"}, inplace=True)

            # Realizar cruce
            resultado = df_deuda.merge(
                pagos_resumen,
                on=["ID_COBRANZA", "PERIODO"],
                how="left"
            )

            resultado["TOTAL_PAGADO"] = resultado["TOTAL_PAGADO"].fillna(0)
            resultado["SALDO_PENDIENTE"] = resultado["DEUDA"] - resultado["TOTAL_PAGADO"]
            resultado["SALDO_PENDIENTE"] = resultado["SALDO_PENDIENTE"].apply(lambda x: max(0, x))

            resultado["ESTADO"] = resultado.apply(
                lambda row: "‚úÖ PAGADO" if row["TOTAL_PAGADO"] >= row["DEUDA"] else "‚è≥ PENDIENTE",
                axis=1
            )

            resultado["PORCENTAJE_PAGADO"] = (resultado["TOTAL_PAGADO"] / resultado["DEUDA"] * 100).round(2)
            resultado["PORCENTAJE_PAGADO"] = resultado["PORCENTAJE_PAGADO"].apply(lambda x: min(100, x))

            st.success("‚úÖ Cruce realizado correctamente")

        except Exception as e:
            st.error(f"‚ùå Error al procesar los archivos: {str(e)}")
            return

    st.markdown("---")

    # ==========================================================
    # DASHBOARD VISUAL - M√âTRICAS PRINCIPALES
    # ==========================================================

    st.markdown("## üìà M√âTRICAS EJECUTIVAS")

    # Calcular m√©tricas principales
    total_cartera = resultado["DEUDA"].sum()
    total_recuperado = resultado["TOTAL_PAGADO"].sum()
    saldo_pendiente = resultado["SALDO_PENDIENTE"].sum()
    porcentaje_recuperacion = (total_recuperado / total_cartera * 100) if total_cartera > 0 else 0
    total_casos = len(resultado)
    casos_pagados = len(resultado[resultado["ESTADO"] == "‚úÖ PAGADO"])
    casos_pendientes = len(resultado[resultado["ESTADO"] == "‚è≥ PENDIENTE"])

    # KPIs principales en cards
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
        st.metric(
            label="üíº CARTERA TOTAL",
            value=f"Bs. {total_cartera:,.2f}",
            delta=f"{total_casos:,} casos"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="metric-card success-card">', unsafe_allow_html=True)
        st.metric(
            label="‚úÖ RECUPERADO",
            value=f"Bs. {total_recuperado:,.2f}",
            delta=f"{porcentaje_recuperacion:.1f}%"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    with col3:
        st.markdown('<div class="metric-card warning-card">', unsafe_allow_html=True)
        st.metric(
            label="‚è≥ PENDIENTE",
            value=f"Bs. {saldo_pendiente:,.2f}",
            delta=f"{casos_pendientes:,} casos"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    with col4:
        delta_color = "normal" if porcentaje_recuperacion >= 70 else "inverse"
        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
        st.metric(
            label="üìä EFECTIVIDAD",
            value=f"{porcentaje_recuperacion:.1f}%",
            delta=f"{casos_pagados:,} pagados"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")

    # ==========================================================
    # FILTROS INTERACTIVOS
    # ==========================================================

    with st.expander("üîç FILTROS Y B√öSQUEDA", expanded=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            periodos_disponibles = ["Todos"] + sorted(resultado["PERIODO"].unique().tolist())
            filtro_periodo = st.selectbox("üìÖ Filtrar por Periodo", periodos_disponibles)
        
        with col2:
            tipos_disponibles = ["Todos"] + sorted(resultado["TIPO"].unique().tolist())
            filtro_tipo = st.selectbox("üè∑Ô∏è Filtrar por Tipo", tipos_disponibles)
        
        with col3:
            estados_disponibles = ["Todos", "‚úÖ PAGADO", "‚è≥ PENDIENTE"]
            filtro_estado = st.selectbox("üìä Filtrar por Estado", estados_disponibles)

    # Aplicar filtros
    resultado_filtrado = resultado.copy()
    
    if filtro_periodo != "Todos":
        resultado_filtrado = resultado_filtrado[resultado_filtrado["PERIODO"] == filtro_periodo]
    
    if filtro_tipo != "Todos":
        resultado_filtrado = resultado_filtrado[resultado_filtrado["TIPO"] == filtro_tipo]
    
    if filtro_estado != "Todos":
        resultado_filtrado = resultado_filtrado[resultado_filtrado["ESTADO"] == filtro_estado]

    # ==========================================================
    # GR√ÅFICOS INTERACTIVOS
    # ==========================================================

    st.markdown("## üìä AN√ÅLISIS VISUAL")

    # Fila 1: Gr√°ficos principales
    col1, col2 = st.columns(2)

    with col1:
        # Gr√°fico de barras: Pagado vs Pendiente
        st.markdown("### üí∞ Comparativa: Recuperado vs Pendiente")
        
        fig_barras = go.Figure()
        
        fig_barras.add_trace(go.Bar(
            name='Recuperado',
            x=['Total'],
            y=[total_recuperado],
            marker_color='#28a745',
            text=[f'Bs. {total_recuperado:,.2f}'],
            textposition='auto',
        ))
        
        fig_barras.add_trace(go.Bar(
            name='Pendiente',
            x=['Total'],
            y=[saldo_pendiente],
            marker_color='#dc3545',
            text=[f'Bs. {saldo_pendiente:,.2f}'],
            textposition='auto',
        ))
        
        fig_barras.update_layout(
            barmode='group',
            height=400,
            showlegend=True,
            hovermode='x unified'
        )
        
        st.plotly_chart(fig_barras, use_container_width=True)

    with col2:
        # Gr√°fico circular: Distribuci√≥n por estado
        st.markdown("### üìà Distribuci√≥n de Casos")
        
        fig_pie = go.Figure(data=[go.Pie(
            labels=['Pagado', 'Pendiente'],
            values=[casos_pagados, casos_pendientes],
            marker=dict(colors=['#28a745', '#ffc107']),
            hole=0.4,
            textinfo='label+percent+value',
            hovertemplate='<b>%{label}</b><br>Casos: %{value}<br>Porcentaje: %{percent}<extra></extra>'
        )])
        
        fig_pie.update_layout(
            height=400,
            showlegend=True,
            annotations=[dict(text=f'{total_casos}<br>Total', x=0.5, y=0.5, font_size=20, showarrow=False)]
        )
        
        st.plotly_chart(fig_pie, use_container_width=True)

    # Fila 2: An√°lisis por periodo y tipo
    col1, col2 = st.columns(2)

    with col1:
        # An√°lisis por periodo
        st.markdown("### üìÖ Evoluci√≥n por Periodo")
        
        periodo_analisis = resultado.groupby("PERIODO").agg({
            "DEUDA": "sum",
            "TOTAL_PAGADO": "sum",
            "SALDO_PENDIENTE": "sum"
        }).reset_index()
        
        fig_periodo = go.Figure()
        
        fig_periodo.add_trace(go.Bar(
            name='Deuda',
            x=periodo_analisis['PERIODO'],
            y=periodo_analisis['DEUDA'],
            marker_color='#667eea'
        ))
        
        fig_periodo.add_trace(go.Bar(
            name='Pagado',
            x=periodo_analisis['PERIODO'],
            y=periodo_analisis['TOTAL_PAGADO'],
            marker_color='#28a745'
        ))
        
        fig_periodo.update_layout(
            barmode='group',
            height=400,
            xaxis_title="Periodo",
            yaxis_title="Monto (Bs.)",
            hovermode='x unified'
        )
        
        st.plotly_chart(fig_periodo, use_container_width=True)

    with col2:
        # An√°lisis por tipo de deuda
        st.markdown("### üè∑Ô∏è Distribuci√≥n por Tipo de Deuda")
        
        tipo_analisis = resultado.groupby("TIPO").agg({
            "DEUDA": "sum",
            "TOTAL_PAGADO": "sum"
        }).reset_index()
        
        tipo_analisis["Pendiente"] = tipo_analisis["DEUDA"] - tipo_analisis["TOTAL_PAGADO"]
        
        fig_tipo = go.Figure()
        
        fig_tipo.add_trace(go.Bar(
            name='Recuperado',
            x=tipo_analisis['TIPO'],
            y=tipo_analisis['TOTAL_PAGADO'],
            marker_color='#28a745'
        ))
        
        fig_tipo.add_trace(go.Bar(
            name='Pendiente',
            x=tipo_analisis['TIPO'],
            y=tipo_analisis['Pendiente'],
            marker_color='#ffc107'
        ))
        
        fig_tipo.update_layout(
            barmode='stack',
            height=400,
            xaxis_title="Tipo de Deuda",
            yaxis_title="Monto (Bs.)",
            hovermode='x unified'
        )
        
        st.plotly_chart(fig_tipo, use_container_width=True)

    # Indicador de efectividad por periodo
    st.markdown("### üéØ Efectividad de Recuperaci√≥n por Periodo")
    
    efectividad_periodo = resultado.groupby("PERIODO").apply(
        lambda x: (x["TOTAL_PAGADO"].sum() / x["DEUDA"].sum() * 100) if x["DEUDA"].sum() > 0 else 0
    ).reset_index()
    efectividad_periodo.columns = ["PERIODO", "EFECTIVIDAD"]
    
    fig_efectividad = go.Figure()
    
    fig_efectividad.add_trace(go.Scatter(
        x=efectividad_periodo['PERIODO'],
        y=efectividad_periodo['EFECTIVIDAD'],
        mode='lines+markers+text',
        line=dict(color='#667eea', width=3),
        marker=dict(size=10, color='#764ba2'),
        text=[f'{val:.1f}%' for val in efectividad_periodo['EFECTIVIDAD']],
        textposition='top center'
    ))
    
    fig_efectividad.add_hline(y=70, line_dash="dash", line_color="green", 
                              annotation_text="Meta: 70%")
    
    fig_efectividad.update_layout(
        height=350,
        xaxis_title="Periodo",
        yaxis_title="Efectividad (%)",
        yaxis_range=[0, 100],
        hovermode='x unified'
    )
    
    st.plotly_chart(fig_efectividad, use_container_width=True)

    st.markdown("---")

    # ==========================================================
    # TABLAS DE AN√ÅLISIS
    # ==========================================================

    st.markdown("## üìã AN√ÅLISIS DETALLADO")

    tab1, tab2, tab3, tab4 = st.tabs([
        "üîù TOP Deudores Pendientes", 
        "üìä Resumen por Periodo",
        "üìÑ Detalle Completo",
        "‚úÖ Casos Pagados"
    ])

    with tab1:
        st.markdown("### üéØ TOP 20 Deudores con Mayor Saldo Pendiente")
        
        pendientes = resultado_filtrado[resultado_filtrado["ESTADO"] == "‚è≥ PENDIENTE"].copy()
        
        if len(pendientes) > 0:
            top_deudores = pendientes.nlargest(20, "SALDO_PENDIENTE")[
                ["ID_COBRANZA", "PERIODO", "TIPO", "DEUDA", "TOTAL_PAGADO", "SALDO_PENDIENTE", "PORCENTAJE_PAGADO"]
            ]
            
            # Formatear para mostrar
            top_deudores_display = top_deudores.copy()
            top_deudores_display["DEUDA"] = top_deudores_display["DEUDA"].apply(lambda x: f"Bs. {x:,.2f}")
            top_deudores_display["TOTAL_PAGADO"] = top_deudores_display["TOTAL_PAGADO"].apply(lambda x: f"Bs. {x:,.2f}")
            top_deudores_display["SALDO_PENDIENTE"] = top_deudores_display["SALDO_PENDIENTE"].apply(lambda x: f"Bs. {x:,.2f}")
            top_deudores_display["PORCENTAJE_PAGADO"] = top_deudores_display["PORCENTAJE_PAGADO"].apply(lambda x: f"{x:.1f}%")
            
            st.dataframe(
                top_deudores_display,
                use_container_width=True,
                height=400
            )
            
            st.metric("üí∞ Saldo Total TOP 20", f"Bs. {top_deudores['SALDO_PENDIENTE'].sum():,.2f}")
        else:
            st.info("‚úÖ ¬°Excelente! No hay casos pendientes con los filtros aplicados.")

    with tab2:
        st.markdown("### üìä Resumen Consolidado por Periodo")
        
        resumen_periodo = resultado_filtrado.groupby("PERIODO").agg({
            "ID_COBRANZA": "count",
            "DEUDA": "sum",
            "TOTAL_PAGADO": "sum",
            "SALDO_PENDIENTE": "sum"
        }).reset_index()
        
        resumen_periodo.columns = ["PERIODO", "CASOS", "DEUDA_TOTAL", "RECUPERADO", "PENDIENTE"]
        resumen_periodo["EFECTIVIDAD_%"] = (resumen_periodo["RECUPERADO"] / resumen_periodo["DEUDA_TOTAL"] * 100).round(2)
        
        # Formatear
        resumen_display = resumen_periodo.copy()
        resumen_display["DEUDA_TOTAL"] = resumen_display["DEUDA_TOTAL"].apply(lambda x: f"Bs. {x:,.2f}")
        resumen_display["RECUPERADO"] = resumen_display["RECUPERADO"].apply(lambda x: f"Bs. {x:,.2f}")
        resumen_display["PENDIENTE"] = resumen_display["PENDIENTE"].apply(lambda x: f"Bs. {x:,.2f}")
        resumen_display["EFECTIVIDAD_%"] = resumen_display["EFECTIVIDAD_%"].apply(lambda x: f"{x:.1f}%")
        
        st.dataframe(
            resumen_display,
            use_container_width=True,
            height=400
        )

    with tab3:
        st.markdown("### üìÑ Detalle Completo de Todos los Casos")
        
        # Preparar datos para mostrar
        resultado_display = resultado_filtrado.copy()
        resultado_display = resultado_display[[
            "ID_COBRANZA", "PERIODO", "TIPO", "DEUDA", "TOTAL_PAGADO", 
            "SALDO_PENDIENTE", "PORCENTAJE_PAGADO", "ESTADO"
        ]]
        
        # Formatear columnas num√©ricas
        for col in ["DEUDA", "TOTAL_PAGADO", "SALDO_PENDIENTE"]:
            resultado_display[col] = resultado_display[col].apply(lambda x: f"Bs. {x:,.2f}")
        resultado_display["PORCENTAJE_PAGADO"] = resultado_display["PORCENTAJE_PAGADO"].apply(lambda x: f"{x:.1f}%")
        
        st.dataframe(
            resultado_display,
            use_container_width=True,
            height=400
        )
        
        st.info(f"üìä Mostrando {len(resultado_filtrado):,} de {len(resultado):,} casos totales")

    with tab4:
        st.markdown("### ‚úÖ Casos Completamente Pagados")
        
        pagados = resultado_filtrado[resultado_filtrado["ESTADO"] == "‚úÖ PAGADO"].copy()
        
        if len(pagados) > 0:
            pagados_display = pagados[[
                "ID_COBRANZA", "PERIODO", "TIPO", "DEUDA", "TOTAL_PAGADO"
            ]].copy()
            
            # Formatear
            pagados_display["DEUDA"] = pagados_display["DEUDA"].apply(lambda x: f"Bs. {x:,.2f}")
            pagados_display["TOTAL_PAGADO"] = pagados_display["TOTAL_PAGADO"].apply(lambda x: f"Bs. {x:,.2f}")
            
            st.dataframe(
                pagados_display,
                use_container_width=True,
                height=400
            )
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("üìã Total Casos Pagados", f"{len(pagados):,}")
            with col2:
                st.metric("üí∞ Monto Total Recuperado", f"Bs. {pagados['TOTAL_PAGADO'].sum():,.2f}")
        else:
            st.warning("‚ö†Ô∏è No hay casos completamente pagados con los filtros aplicados.")

    st.markdown("---")

    # ==========================================================
    # EXPORTACI√ìN A EXCEL
    # ==========================================================

    st.markdown("## üì• EXPORTAR RESULTADOS")

    col1, col2, col3 = st.columns(3)

    with col1:
        # Exportar todo
        if st.button("üìä Exportar Dashboard Completo", use_container_width=True):
            with st.spinner("Generando reporte Excel..."):
                buffer = generar_reporte_excel(resultado, resultado_filtrado)
                
                st.download_button(
                    label="‚¨áÔ∏è Descargar Reporte Completo",
                    data=buffer,
                    file_name=f"Reporte_Cobranza_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    with col2:
        # Exportar solo pendientes
        if st.button("‚è≥ Exportar Solo Pendientes", use_container_width=True):
            pendientes = resultado_filtrado[resultado_filtrado["ESTADO"] == "‚è≥ PENDIENTE"]
            
            if len(pendientes) > 0:
                buffer = exportar_simple(pendientes, "Casos Pendientes")
                
                st.download_button(
                    label="‚¨áÔ∏è Descargar Pendientes",
                    data=buffer,
                    file_name=f"Pendientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.info("‚úÖ No hay casos pendientes")

    with col3:
        # Exportar TOP deudores
        if st.button("üéØ Exportar TOP 50 Deudores", use_container_width=True):
            pendientes = resultado_filtrado[resultado_filtrado["ESTADO"] == "‚è≥ PENDIENTE"]
            
            if len(pendientes) > 0:
                top_50 = pendientes.nlargest(50, "SALDO_PENDIENTE")
                buffer = exportar_simple(top_50, "TOP 50 Deudores")
                
                st.download_button(
                    label="‚¨áÔ∏è Descargar TOP 50",
                    data=buffer,
                    file_name=f"TOP_Deudores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.info("‚úÖ No hay casos pendientes")


def generar_reporte_excel(resultado_completo, resultado_filtrado):
    """Genera un reporte Excel profesional con m√∫ltiples hojas"""
    
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        
        # Hoja 1: Resumen Ejecutivo
        resumen_data = {
            "M√âTRICA": [
                "Total Cartera",
                "Total Recuperado",
                "Saldo Pendiente",
                "% Efectividad",
                "Total Casos",
                "Casos Pagados",
                "Casos Pendientes"
            ],
            "VALOR": [
                f"Bs. {resultado_completo['DEUDA'].sum():,.2f}",
                f"Bs. {resultado_completo['TOTAL_PAGADO'].sum():,.2f}",
                f"Bs. {resultado_completo['SALDO_PENDIENTE'].sum():,.2f}",
                f"{(resultado_completo['TOTAL_PAGADO'].sum() / resultado_completo['DEUDA'].sum() * 100):.2f}%",
                len(resultado_completo),
                len(resultado_completo[resultado_completo['ESTADO'] == '‚úÖ PAGADO']),
                len(resultado_completo[resultado_completo['ESTADO'] == '‚è≥ PENDIENTE'])
            ]
        }
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)
        
        # Hoja 2: Detalle Completo
        resultado_completo.to_excel(writer, sheet_name='Detalle Completo', index=False)
        
        # Hoja 3: Solo Pendientes
        pendientes = resultado_completo[resultado_completo['ESTADO'] == '‚è≥ PENDIENTE']
        pendientes.to_excel(writer, sheet_name='Pendientes', index=False)
        
        # Hoja 4: TOP 50 Deudores
        if len(pendientes) > 0:
            top_50 = pendientes.nlargest(50, 'SALDO_PENDIENTE')
            top_50.to_excel(writer, sheet_name='TOP 50 Deudores', index=False)
        
        # Hoja 5: Resumen por Periodo
        resumen_periodo = resultado_completo.groupby('PERIODO').agg({
            'ID_COBRANZA': 'count',
            'DEUDA': 'sum',
            'TOTAL_PAGADO': 'sum',
            'SALDO_PENDIENTE': 'sum'
        }).reset_index()
        resumen_periodo.columns = ['PERIODO', 'CASOS', 'DEUDA_TOTAL', 'RECUPERADO', 'PENDIENTE']
        resumen_periodo['EFECTIVIDAD_%'] = (resumen_periodo['RECUPERADO'] / resumen_periodo['DEUDA_TOTAL'] * 100).round(2)
        resumen_periodo.to_excel(writer, sheet_name='Por Periodo', index=False)
        
        # Aplicar formato
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Formato de encabezados
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    buffer.seek(0)
    return buffer


def exportar_simple(df, nombre_hoja):
    """Exporta un DataFrame simple a Excel"""
    
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nombre_hoja, index=False)
        
        # Formato b√°sico
        workbook = writer.book
        worksheet = workbook[nombre_hoja]
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    buffer.seek(0)
    return buffer


# ==========================================================
# MODULO 2 - GENERADOR DE SMS (SIN CAMBIOS)
# ==========================================================

def modulo_sms():

    st.title("üì≤ GENERADOR PROFESIONAL DE SMS")

    def limpiar_columnas(df):
        df.columns = df.columns.str.strip().str.upper().str.replace(" ", "_")
        return df

    archivo_suscriptor = st.file_uploader("üìÇ Cargar BASE POR SUSCRIPTOR", type=["xlsx"])
    archivo_pagos = st.file_uploader("üíµ Cargar BASE DE PAGOS", type=["xlsx"])

    if not archivo_suscriptor or not archivo_pagos:
        return

    df_suscriptor = limpiar_columnas(pd.read_excel(archivo_suscriptor))
    df_pagos = limpiar_columnas(pd.read_excel(archivo_pagos))

    df_suscriptor["CODIGO"] = df_suscriptor["CODIGO"].astype(str)
    df_suscriptor["MONTO"] = pd.to_numeric(df_suscriptor["MONTO"], errors="coerce").fillna(0)
    df_pagos["ID_COBRANZA"] = df_pagos["ID_COBRANZA"].astype(str)
    df_pagos["IMPORTE"] = pd.to_numeric(df_pagos["IMPORTE"], errors="coerce").fillna(0)

    pagos_totales = df_pagos.groupby("ID_COBRANZA")["IMPORTE"].sum().reset_index()
    pagos_totales.rename(columns={"IMPORTE": "TOTAL_PAGADO"}, inplace=True)

    df_final = df_suscriptor.merge(
        pagos_totales,
        left_on="CODIGO",
        right_on="ID_COBRANZA",
        how="left"
    )

    df_final["TOTAL_PAGADO"] = df_final["TOTAL_PAGADO"].fillna(0)
    df_final = df_final[df_final["TOTAL_PAGADO"] < df_final["MONTO"]]

    columnas_exportar = ["NUMERO", "NOMBRE", "FECHA", "CODIGO", "MONTO"]
    df_export = df_final[columnas_exportar].copy()

    st.subheader("Vista previa final")
    st.dataframe(df_export)

    partes = st.number_input("Cantidad de archivos CSV", min_value=1, value=1)
    prefijo = st.text_input("Prefijo archivos", value="SMS")

    if st.button("Generar CSV"):

        if df_export.empty:
            st.warning("No existen registros.")
            return

        tama√±o = len(df_export) // partes + 1

        for i in range(partes):
            inicio = i * tama√±o
            fin = inicio + tama√±o
            df_parte = df_export.iloc[inicio:fin]

            if df_parte.empty:
                continue

            csv = df_parte.to_csv(
                index=False,
                sep=";",
                encoding="utf-8-sig"
            )

            st.download_button(
                label=f"Descargar {prefijo}_{i+1}.csv",
                data=csv,
                file_name=f"{prefijo}_{i+1}.csv",
                mime="text/csv"
            )


# ==========================================================
# EJECUCI√ìN
# ==========================================================

if menu == "üìä Dashboard Cruce Deuda vs Pagos":
    modulo_cruce()

elif menu == "üì≤ GENERADOR DE SMS":
    modulo_sms()

elif menu == "üöß M√≥dulo Hist√≥rico (En Desarrollo)":
    st.title("üìà Hist√≥rico")
    st.info("Aqu√≠ construiremos el dashboard acumulado mensual.")
